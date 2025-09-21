# Streamlit Excel Validation App
# Validates uploaded Excel files against specified rules and exports a styled Excel with highlights and messages.

import io
from io import BytesIO
from typing import Dict, List, Tuple, Any

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Alignment

# -----------------------------
# Configuration
# -----------------------------
EXPECTED_COLUMNS = [
    "macroid",
    "asset_type",
    "asset_name",
    "final_value",
    "asset_usage_id",
    "value_base",
    "inspection_date",
    "production_capacity",
    "production_capacity_measuring_unit",
    "owner_name",
    "product_type",
    "market_approach",
    "market_approach_value",
    "cost_approach",
    "cost_approach_value",
    "country",
    "region",
    "city",
]

MANDATORY_FIELDS = [
    # Note: cost_approach and cost_approach_value were NOT listed as mandatory by the user
    "asset_type",
    "asset_name",
    "asset_usage_id",
    "value_base",
    "inspection_date",
    "final_value",
    "production_capacity",
    "production_capacity_measuring_unit",
    "owner_name",
    "product_type",
    "market_approach",
    "market_approach_value",
    "country",
    "region",
    "city",
]

# Valid ranges/maps
ASSET_USAGE_MIN, ASSET_USAGE_MAX = 38, 56  # inclusive
VALUE_BASE_MIN, VALUE_BASE_MAX = 1, 9      # inclusive
MARKET_APPROACH_ALLOWED = {0, 1, 2}

# Colors (ARGB) for fills
FILL_RED = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")     # Errors - critical
FILL_YELLOW = PatternFill(start_color="FFDE21", end_color="FFDE21", fill_type="solid")  # Missing mandatory/problem fields
FILL_ORANGE = PatternFill(start_color="FFFFE4B5", end_color="FFFFE4B5", fill_type="solid")  # Date issues

# -----------------------------
# Utility functions
# -----------------------------

def is_empty(value: Any) -> bool:
    """Return True if value should be considered empty.
    '0' and 'N/A' (case-insensitive) are NOT considered empty.
    """
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    s = str(value).strip()
    if s == "":
        return True
    if s.lower() == "n/a":
        # Considered a valid placeholder, not empty
        return False
    if s == "0":
        # 0 is valid for numeric/flag fields
        return False
    return False if s else True


def to_int(value: Any) -> Tuple[bool, int | None]:
    """Try to parse int without decimals. Returns (ok, int_value)."""
    if is_empty(value):
        return False, None
    try:
        # Allow numeric strings like '97000' but not '97000.5'
        s = str(value).strip()
        if s.lower().endswith(".0"):
            s = s[:-2]
        if "." in s:
            return False, None
        return True, int(s)
    except Exception:
        return False, None


def to_float(value: Any) -> Tuple[bool, float | None]:
    """Parse value to float. Returns (ok, float_value)."""
    if is_empty(value):
        return False, None
    try:
        return True, float(str(value).strip())
    except Exception:
        return False, None


def parse_and_format_date(value: Any) -> Tuple[bool, str | None]:
    """Try to parse a date and return it formatted as dd-mm-YYYY. If not parseable, return (False, None)."""
    if is_empty(value):
        return False, None
    s = str(value).strip()
    # Try multiple common formats, dayfirst where applicable
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="raise")
        return True, dt.strftime("%d-%m-%Y")
    except Exception:
        return False, None


# -----------------------------
# Validation functions
# -----------------------------

def check_missing_columns(df: pd.DataFrame) -> List[str]:
    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    return missing


def validate_final_value_only(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[Tuple[int, str], PatternFill], List[str]]:
    """Validate only final_value emptiness and integer-ness.
    Writes the validation message directly into the offending cell and colors it yellow.
    Returns updated df, cell highlights, and summary lines.
    """
    df = df.copy()

    highlights: Dict[Tuple[int, str], PatternFill] = {}
    summary: List[str] = []
    issues = 0

    if "final_value" in df.columns:
        for idx, val in df["final_value"].items():
            message = None
            if is_empty(val):
                message = "final_value is mandatory and cannot be empty"
            else:
                ok, _intval = to_int(val)
                if not ok:
                    message = "Final value must be a non-decimal integer"
            if message:
                issues += 1
                highlights[(idx, "final_value")] = FILL_YELLOW
                df.at[idx, "final_value"] = append_message(df.at[idx, "final_value"], message)

    summary.append(f"Final Value issues: {issues}")
    return df, highlights, summary


def validate_mandatory_only(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[Tuple[int, str], PatternFill], List[str]]:
    df = df.copy()

    highlights: Dict[Tuple[int, str], PatternFill] = {}
    summary: List[str] = []

    missing_count = 0

    for col in MANDATORY_FIELDS:
        if col not in df.columns:
            continue  # Column presence is checked elsewhere; here we only mark empty values
        for idx, val in df[col].items():
            if is_empty(val):
                # Special case: market_approach can be empty => treat as 0 (allowed) - not flagged here
                if col == "market_approach":
                    continue
                # For market_approach_value: allow empty if approach is 0/empty
                if col == "market_approach_value":
                    try:
                        approach_raw = df.at[idx, "market_approach"] if "market_approach" in df.columns else ""
                        approach = 0 if is_empty(approach_raw) else int(float(str(approach_raw).strip()))
                    except Exception:
                        approach = None
                    if approach in (0, None):
                        continue  # ok to be empty
                missing_count += 1
                highlights[(idx, col)] = FILL_YELLOW
                df.at[idx, col] = append_message(df.at[idx, col], "This mandatory field is empty")

    summary.append(f"Missing mandatory values: {missing_count}")
    return df, highlights, summary


def validate_dates_only(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[Tuple[int, str], PatternFill], List[str]]:
    df = df.copy()

    highlights: Dict[Tuple[int, str], PatternFill] = {}
    summary: List[str] = []

    invalid_count = 0
    auto_fixed = 0

    if "inspection_date" in df.columns:
        for idx, val in df["inspection_date"].items():
            if is_empty(val):
                highlights[(idx, "inspection_date")] = FILL_YELLOW
                df.at[idx, "inspection_date"] = append_message(df.at[idx, "inspection_date"], "Date must be in dd-mm-YYYY format")
                invalid_count += 1
                continue
            ok, formatted = parse_and_format_date(val)
            if ok and formatted:
                # Auto-fix to desired format
                df.at[idx, "inspection_date"] = formatted
                auto_fixed += 1
            else:
                highlights[(idx, "inspection_date")] = FILL_YELLOW
                df.at[idx, "inspection_date"] = append_message(df.at[idx, "inspection_date"], "Date must be in dd-mm-YYYY format")
                invalid_count += 1
    else:
        summary.append("Column 'inspection_date' is missing")

    summary.append(f"Invalid dates: {invalid_count}")
    if auto_fixed:
        summary.append(f"Dates auto-formatted: {auto_fixed}")
    return df, highlights, summary


def validate_all(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[Tuple[int, str], PatternFill], List[str]]:
    """Run all validations: mandatory emptiness, final value integer, date format, and range checks.
    Messages are written directly inside the invalid cells and colored accordingly.
    """
    df = df.copy()

    highlights: Dict[Tuple[int, str], PatternFill] = {}
    summary: List[str] = []

    # 1) Mandatory non-empty (with allowed exceptions)
    df, hl_mand, sum_mand = validate_mandatory_only(df)
    highlights.update(hl_mand)
    summary.extend(sum_mand)

    # 2) Final value integer
    df, hl_final, sum_final = validate_final_value_only(df)
    highlights.update(hl_final)
    summary.extend(sum_final)

    # 3) Dates
    df, hl_dates, sum_dates = validate_dates_only(df)
    highlights.update(hl_dates)
    summary.extend(sum_dates)

    # 4) Additional numeric/range checks
    extra_issues = 0
    # asset_usage_id: integer 38..56
    if "asset_usage_id" in df.columns:
        for idx, val in df["asset_usage_id"].items():
            if is_empty(val):
                continue  # already handled by mandatory check
            ok, intval = to_int(val)
            if not ok or intval is None or not (ASSET_USAGE_MIN <= intval <= ASSET_USAGE_MAX):
                highlights[(idx, "asset_usage_id")] = FILL_YELLOW
                df.at[idx, "asset_usage_id"] = append_message(df.at[idx, "asset_usage_id"], f"asset_usage_id must be in [{ASSET_USAGE_MIN}-{ASSET_USAGE_MAX}]")
                extra_issues += 1

    # value_base: integer 1..9
    if "value_base" in df.columns:
        for idx, val in df["value_base"].items():
            if is_empty(val):
                continue
            ok, intval = to_int(val)
            if not ok or intval is None or not (VALUE_BASE_MIN <= intval <= VALUE_BASE_MAX):
                highlights[(idx, "value_base")] = FILL_YELLOW
                df.at[idx, "value_base"] = append_message(df.at[idx, "value_base"], f"value_base must be in [{VALUE_BASE_MIN}-{VALUE_BASE_MAX}]")
                extra_issues += 1

    # market_approach: 0,1,2 (empty treated as 0)
    if "market_approach" in df.columns:
        for idx, val in df["market_approach"].items():
            if is_empty(val):
                # treat as 0
                continue
            ok, fval = to_float(val)
            if not ok or fval is None or int(fval) not in MARKET_APPROACH_ALLOWED:
                highlights[(idx, "market_approach")] = FILL_YELLOW
                df.at[idx, "market_approach"] = append_message(df.at[idx, "market_approach"], "market_approach must be 0, 1, or 2")
                extra_issues += 1

    # market_approach_value: must be provided and numeric if approach in {1,2}; allowed empty if approach 0/empty
    if "market_approach_value" in df.columns:
        for idx, val in df["market_approach_value"].items():
            approach_raw = df.at[idx, "market_approach"] if "market_approach" in df.columns else ""
            approach_empty = is_empty(approach_raw)
            approach = 0
            if not approach_empty:
                try:
                    approach = int(float(str(approach_raw).strip()))
                except Exception:
                    approach = None
            if approach in (1, 2):
                ok, fval = to_float(val)
                if not ok or fval is None:
                    highlights[(idx, "market_approach_value")] = FILL_YELLOW
                    df.at[idx, "market_approach_value"] = append_message(df.at[idx, "market_approach_value"], "Must be a number when approach is 1 or 2")
                    extra_issues += 1

    # production_capacity: if provided, must be non-negative number (it's mandatory; emptiness handled already)
    if "production_capacity" in df.columns:
        for idx, val in df["production_capacity"].items():
            if is_empty(val):
                continue
            ok, fval = to_float(val)
            if not ok or fval is None or fval < 0:
                highlights[(idx, "production_capacity")] = FILL_YELLOW
                df.at[idx, "production_capacity"] = append_message(df.at[idx, "production_capacity"], "Must be a non-negative number")
                extra_issues += 1

    summary.append(f"Additional rule violations: {extra_issues}")

    return df, highlights, summary


# -----------------------------
# Styling export
# -----------------------------

def append_message(existing: Any, new_msg: str) -> str:
    s = str(existing).strip()
    if not s or s.lower() == "nan":
        return new_msg
    return f"{s} | {new_msg}"


def export_with_highlights(df: pd.DataFrame, highlights: Dict[Tuple[int, str], PatternFill]) -> bytes:
    """Export DataFrame to Excel with openpyxl and apply cell highlights.
    highlights keys are (row_index_in_df, column_name) -> PatternFill.
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write the DataFrame
        df.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.sheets["Data"]

        # Build column name -> Excel column index map
        headers = [cell.value for cell in ws[1]]
        col_index_by_name = {name: i + 1 for i, name in enumerate(headers)}  # 1-based

        # Apply RTL alignment for text cells to better handle Arabic
        right_align = Alignment(horizontal="right")

        # Apply highlights
        for (row_idx_df, col_name), fill in highlights.items():
            if col_name not in col_index_by_name:
                continue
            row_xl = row_idx_df + 2  # +1 for header, +1 to convert 0-based to 1-based
            col_xl = col_index_by_name[col_name]
            cell = ws.cell(row=row_xl, column=col_xl)
            cell.fill = fill
            # Right-align for readability with RTL
            cell.alignment = right_align

        # Also right-align all header cells for RTL readability
        for cell in ws[1]:
            cell.alignment = right_align

    return output.getvalue()


# -----------------------------
# Streamlit UI
# -----------------------------

def main():
    st.set_page_config(page_title="Excel Validator", layout="wide")

    # Professional UI with Bootstrap, custom colors, and animation
    st.markdown(
        """
        <link rel='stylesheet' href='https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.rtl.min.css'>
        <link rel='stylesheet' href='https://cdnjs.cloudflare.com/ajax/libs/animate.css/4.1.1/animate.min.css'/>
        <style>
        html, body, [class^="css"]  {
            direction: rtl;
            background: #f7f8fa !important;
            color: #111 !important;
        }
        .stApp {
            background: linear-gradient(135deg, #fffbe6 0%, #f7f8fa 100%) !important;
            color: #111 !important;
        }
        .main-title, .main-desc, .stSubheader, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stDownloadButton button, .stButton > button, .stAlert, .stInfo, .stSuccess, .stError, .stDataFrame, .block-container, .stFileUploader, .stMarkdown p, .stMarkdown ul, .stMarkdown li, .stMarkdown span, .stMarkdown div {
            color: #111 !important;
        }
        .stAlert, .stInfo, .stSuccess, .stError, .stNotification, .stNotificationContent, .stNotificationIcon, .stNotificationText, .stNotificationBody, .stNotificationMessage, .stNotificationContent *, .stNotification *, .stInfo *, .stError *, .stSuccess * {
            color: #111 !important;
        }
        /* Keep background and other styles as before */
        .main-title {
            font-family: 'Cairo', 'Tajawal', Arial, sans-serif;
            font-weight: 900;
            font-size: 2.5em;
            text-shadow: 0 2px 8px #ffe06633;
            margin-bottom: 0.5em;
            background: linear-gradient(90deg, #FFDE21 0%, #FFD700 100%);
            border-radius: 18px;
            padding: 0.5em 1em;
            box-shadow: 0 4px 32px #ffe06633;
            display: inline-block;
            animation: fadeInDown 1s;
        }
        .main-desc {
            font-size: 1.2em;
            margin-bottom: 1.5em;
            animation: fadeIn 1.5s;
        }
        .stDownloadButton button, .stButton > button {
            background: linear-gradient(90deg, #21a366 0%, #43c97f 100%);
            color: #fff !important;
            border-radius: 12px;
            font-weight: bold;
            border: none;
            box-shadow: 0 4px 16px #21a36633;
            padding: 0.75em 2em;
            font-size: 1.1em;
            margin-bottom: 12px;
            transition: background 0.2s, box-shadow 0.2s, transform 0.2s;
            letter-spacing: 0.5px;
            animation: none !important;
        }
        .stDownloadButton button:hover, .stButton > button:hover {
            background: linear-gradient(90deg, #43c97f 0%, #21a366 100%);
            color: #fff !important;
            box-shadow: 0 6px 24px #21a36655;
            transform: scale(1.04);
            animation: none !important;
        }
        /* Remove animation from buttons */
        .stDownloadButton button, .stButton > button {
            animation: none !important;
        }
        /* ...existing code for buttons, alerts, etc... */
        .stSuccess {
            background: #111 !important;
            color: #21a366 !important;
            border-radius: 12px !important;
            font-weight: 700 !important;
            font-size: 1.5em !important;
            box-shadow: 0 2px 8px #b6e6b655 !important;
        }
        .stError {
            background: #111 !important;
            color: #ff4b4b !important;
            border-radius: 12px !important;
            font-weight: 700 !important;
            font-size: 1.5em !important;
            box-shadow: 0 2px 8px #ffcccc55 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown('<div class="main-title animate__animated animate__fadeInDown">أداة التحقق من ملف إكسل - Excel Validation Tool</div>', unsafe_allow_html=True)
    st.markdown('<div class="main-desc animate__animated animate__fadeIn">قم برفع ملف Excel بصيغة xlsx للتحقق من الحقول والقيم وفقًا للقواعد المحددة، ثم تنزيل ملف معدل مع إبراز الأخطاء.</div>', unsafe_allow_html=True)

    uploaded_file = st.file_uploader("Upload Excel file (xlsx)", type=["xlsx"])

    # Custom message background below upload
    st.markdown(
        """
        <div style='background: linear-gradient(90deg, #FFDE21 0%, #FFD700 100%); border-radius: 14px; padding: 1.2em; margin-bottom: 1.5em; color: #222; font-size: 1.15em; font-weight: 600; box-shadow: 0 2px 12px #ffe06655;'>
        بعد رفع الملف، سيتم التحقق من الحقول والقيم تلقائيًا. يمكنك تنزيل ملف النتائج بعد التحقق.
        </div>
        """,
        unsafe_allow_html=True
    )

    # Custom CSS for Streamlit's success and error messages
    st.markdown(
        """
        <style>
        .stAlert, .stError {
            background: linear-gradient(90deg, #ffb3b3 0%, #ffdddd 100%) !important;
            color: #a80000 !important;
            border-radius: 12px !important;
            font-weight: 600 !important;
            font-size: 1.1em !important;
            box-shadow: 0 2px 8px #ffcccc55 !important;
        }
        .stSuccess {
            background: linear-gradient(90deg, #d4fc79 0%, #96e6a1 100%) !important;
            color: #176a00 !important;
            border-radius: 12px !important;
            font-weight: 600 !important;
            font-size: 1.1em !important;
            box-shadow: 0 2px 8px #b6e6b655 !important;
        }
        .stSubheader, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 {
            color: #222 !important;
        }
        .stMarkdown ul, .stMarkdown li, .stMarkdown p, .stMarkdown span, .stMarkdown div {
            color: #111 !important;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    if uploaded_file is None:
        st.markdown(
            """
            <div style='background: #f7f8fa; color: #111; border-radius: 10px; padding: 1em; font-size: 1.1em; font-weight: 600; border: 1px solid #eee; margin-bottom: 1em;'>
            Please upload an Excel file to begin.
            </div>
            """,
            unsafe_allow_html=True
        )
        return

    try:
        # Read as string to preserve formatting; we'll parse as needed
        df = pd.read_excel(uploaded_file, dtype=str)
    except Exception as e:
        st.error(f"Could not read the Excel file: {e}")
        return

    # Ensure expected columns exist
    missing_cols = check_missing_columns(df)
    if missing_cols:
        st.error(
            "The uploaded file is missing required columns: " + ", ".join(missing_cols)
        )
        st.stop()

    # Place action buttons
    col1, col2, col3, col4 = st.columns(4)
    do_final = col1.button("Check Final Value")
    do_mand = col2.button("Check Mandatory Fields")
    do_date = col3.button("Check Date Format")
    do_all = col4.button("Check All")

    # Add button to sum asset values (final_value column)
    col_sum, _ = st.columns([1, 3])
    if col_sum.button("Sum Asset Values"):
        if "final_value" in df.columns:
            total = pd.to_numeric(df["final_value"], errors="coerce").sum()
            st.markdown(f"""
            <div style='color: #111; font-weight: bold; font-size: 2em; background: #fff; border-radius: 14px; padding: 0.7em 1em; margin-bottom: 1em; box-shadow: 0 2px 8px #b6e6b655;'>
            Total Asset Value (sum of final_value): {total:,.2f}
            </div>
            """, unsafe_allow_html=True)
        else:
            st.error("Column 'final_value' not found in the uploaded file.")

    # Default outputs
    out_df = df.copy()
    highlights: Dict[Tuple[int, str], PatternFill] = {}
    summary: List[str] = []

    # Execute selected validation
    if do_final:
        out_df, highlights, summary = validate_final_value_only(df.copy())
    elif do_mand:
        out_df, highlights, summary = validate_mandatory_only(df.copy())
    elif do_date:
        out_df, highlights, summary = validate_dates_only(df.copy())
    elif do_all:
        out_df, highlights, summary = validate_all(df.copy())

    if do_final or do_mand or do_date or do_all:
        total_issues = sum(int(s.split(":")[-1].strip()) for s in summary if ":" in s and s.split(":")[0] in [
            "Final Value issues", "Missing mandatory values", "Invalid dates", "Additional rule violations"
        ])

        # Show summary
        st.subheader("Summary")
        for line in summary:
            st.write("- " + line)

        if total_issues == 0:
            st.markdown(
                """
                <div style='color: #21a366; font-weight: bold; font-size: 1.1em; border-radius: 12px; padding: 0.7em 1em;'>
                No issues found. Your file looks good!
                </div>
                """,
                unsafe_allow_html=True
            )

        # Export with highlights and provide download
        try:
            xlsx_bytes = export_with_highlights(out_df, highlights)
            st.download_button(
                label="Download Validated Excel",
                data=xlsx_bytes,
                file_name="validated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Error while generating the Excel file: {e}")

        # Show a styled preview table (Excel-like)
        st.subheader("Preview (all rows)")

        def highlight_excel(row):
            row_styles = []
            for col in out_df.columns:
                key = (row.name, col)
                if key in highlights:
                    row_styles.append('background-color: #FFDE21; color: #111; border: 1px solid #bdbdbd;')
                else:
                    row_styles.append('background-color: #fff; color: #111; border: 1px solid #bdbdbd;')
            return row_styles

        styled = (
            out_df.style
            .apply(highlight_excel, axis=1)
            .set_table_styles([
                {'selector': 'th', 'props': [
                    ('background-color', '#21a366'),
                    ('color', '#fff'),
                    ('font-weight', 'bold'),
                    ('border', '1px solid #bdbdbd'),
                    ('text-align', 'center')
                ]},
                {'selector': 'td', 'props': [
                    ('border', '1px solid #bdbdbd'),
                    ('text-align', 'center')
                ]},
                {'selector': 'table', 'props': [
                    ('background-color', '#fff'),
                    ('border-collapse', 'collapse'),
                    ('font-family', 'Segoe UI, Arial, sans-serif'),
                    ('font-size', '1em')
                ]}
            ])
            .set_properties(**{'text-align': 'center'})
        )
        st.markdown(
            styled.to_html(escape=False, index=False),
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            """
            <div style='background: #f7f8fa; color: #111; border-radius: 10px; padding: 1em; font-size: 1.1em; font-weight: 600; border: 1px solid #eee; margin-bottom: 1em;'>
            Choose a check to run using the buttons above.
            </div>
            """,
            unsafe_allow_html=True
        )


if __name__ == "__main__":
    main()